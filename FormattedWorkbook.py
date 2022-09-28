import sys

import loguru
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from pandas import DataFrame


def fill_cell_names():
    """
        Заполнение словаря для обращения к ячейкам Excel 1:A, 2:B,... 27:AA, 28:AB и так далее до ZZZ

        :return Dictionary:
        """
    _count = 1
    _cell_names = {}

    for _i in range(65, 91):
        _cell_names[_count] = chr(_i)
        _count += 1
    for _i in range(65, 91):
        for _j in range(65, 91):
            _cell_names[_count] = chr(_i) + chr(_j)
            _count += 1
    for _i in range(65, 91):
        for _j in range(65, 91):
            for _k in range(65, 91):
                _cell_names[_count] = chr(_i) + chr(_j) + chr(_k)
                _count += 1
    return _cell_names


def adjust_columns_width(_dataframe):
    # Форматирование ширины полей отчётной таблицы
    for _col in _dataframe.columns:
        _max_length = 0
        _column = get_column_letter(_col[0].column)  # Get the column name
        for _cell in _col:
            if _cell.coordinate in _dataframe.merged_cells:  # not check merge_cells
                continue
            try:  # Necessary to avoid error on empty cells
                if len(str(_cell.value)) > _max_length:
                    _max_length = len(str(_cell.value))
            except Exception as e:
                loguru.logger.debug(f"Empty cell. Error text: {e}")
                pass
        _adjusted_width = _max_length + 3
        _dataframe.column_dimensions[_column].width = _adjusted_width
    return _dataframe


class FormattedWorkbook(Workbook):
    def __init__(self, logging_level='ERROR', table_style='TableStyleMedium2'):
        super().__init__()
        self.logging_level = logging_level
        self.logger = loguru.logger
        self.table_style = table_style
        self.excel_cell_names = fill_cell_names()
        self.ws = self.active

    def excel_format_table(self, df: DataFrame, save_sheet_name: str, save_table_name: str):
        """ Метод обеспечивает форматирование листа Excel с таблицей."""
        self.logger.remove()
        self.logger.add(sys.stdout, level=self.logging_level)
        self.logger.info(f'Создаем лист "{save_sheet_name}"')
        self.ws = self.create_sheet(title=f'{save_sheet_name}')
        self.logger.info(f'Заполняем лист "{save_sheet_name}" данными')
        for row in dataframe_to_rows(df, index=False, header=True):
            self.ws.append(row)
        self.logger.info(f'Форматирует таблицу "{save_table_name}"')
        self.logger.debug(f'Таблица для форматирования: A1:{self.excel_cell_names[len(df.columns)]}{len(df) + 1}')
        tab = Table(displayName=f'{save_table_name}',
                    ref=f'A1:{self.excel_cell_names[len(df.columns)]}{len(df) + 1}')
        tab.tableStyleInfo = TableStyleInfo(name=self.table_style, showRowStripes=True, showColumnStripes=True)
        self.logger.info(f'Добавляем таблицу "{save_table_name}" на лист "{save_sheet_name}"')
        self.ws.add_table(tab)
        self.logger.info(f'Выравниваем поля по размеру в таблице "{save_table_name}"')
        self.ws = adjust_columns_width(self.ws)
